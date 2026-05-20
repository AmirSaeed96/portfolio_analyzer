import os
import shutil
from datetime import datetime, timedelta

import openpyxl
import pandas as pd
import yfinance as yf

# Docker supplies EXCEL_PATH via env var.
# On Streamlit Cloud there's no env var, so we copy the template from the repo
# into /tmp/ (writable, persists for the session) and work from there.
_ENV_PATH = os.environ.get("EXCEL_PATH")
_TEMPLATE  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "ETF_DETAILS_daily_update_interface.xlsx")
_TMP_PATH  = "/tmp/ETF_DETAILS_daily_update_interface.xlsx"

if _ENV_PATH:
    EXCEL_PATH = _ENV_PATH
else:
    EXCEL_PATH = _TMP_PATH
    if not os.path.exists(_TMP_PATH) and os.path.exists(_TEMPLATE):
        shutil.copy(_TEMPLATE, _TMP_PATH)

# These are money market funds — not exchange-traded, stable $1.00 NAV
MONEY_MARKETS = {"FXFXX", "FDRXX", "NOSXX", "FGXXX"}

# Excel symbol → yfinance symbol where they differ
TICKER_MAP = {
    "BRKB": "BRK-B",
    "BRK.B": "BRK-B",
}


def get_tickers(excel_path=None):
    """Return list of (ticker, name) from the Daily Update sheet."""
    path = excel_path or EXCEL_PATH
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Daily Update"]
    tickers = []
    seen = set()
    for row in ws.iter_rows(min_row=6, max_col=2, values_only=True):
        ticker = row[0]
        name = row[1] if len(row) > 1 else ticker
        if ticker and isinstance(ticker, str) and " " not in ticker.strip() and ticker.strip() not in seen:
            sym = ticker.strip()
            name_str = str(name).strip() if name is not None else sym
            tickers.append((sym, name_str))
            seen.add(sym)
    wb.close()
    return tickers


def _price_at(close: pd.Series, target_date) -> float | None:
    """Closing price on the nearest trading day on or before target_date."""
    ts = pd.Timestamp(target_date)
    if close.index.tz is not None:
        ts = ts.tz_localize(close.index.tz)
    subset = close[close.index <= ts]
    return float(subset.iloc[-1]) if not subset.empty else None


def _pct(close: pd.Series, *, days_back: int | None = None, cal_days: int | None = None) -> float | None:
    """Decimal % change (e.g. 0.05 = +5%). Uses trading-day offset or calendar-day lookback."""
    if close.empty:
        return None
    current = float(close.iloc[-1])
    if days_back is not None:
        if len(close) <= days_back:
            return None
        past = float(close.iloc[-(days_back + 1)])
    else:
        latest = close.index[-1]
        target = (latest - pd.Timedelta(days=cal_days)).date()
        past = _price_at(close, target)
    if past is None or past == 0:
        return None
    return (current / past) - 1


def fetch_one(symbol: str) -> tuple[str, dict]:
    """Fetch 5-year history for one ticker and compute all metrics."""
    yf_sym = TICKER_MAP.get(symbol, symbol)
    try:
        hist = yf.Ticker(yf_sym).history(period="5y", auto_adjust=True)
        if hist.empty:
            return symbol, {"error": "No data returned"}
        close = hist["Close"].dropna()
        volume = hist["Volume"].dropna()
        if close.empty:
            return symbol, {"error": "No closing price data"}
        return symbol, {
            "price": round(float(close.iloc[-1]), 4),
            "volume": int(volume.iloc[-1]) if not volume.empty else None,
            "1d": _pct(close, days_back=1),
            "5d": _pct(close, days_back=5),
            "1mo": _pct(close, cal_days=30),
            "1yr": _pct(close, cal_days=365),
            "2yr": _pct(close, cal_days=730),
            "5yr": _pct(close, cal_days=1825),
        }
    except Exception as exc:
        return symbol, {"error": str(exc)}


def write_to_sheet1(results: dict, excel_path=None):
    """Write fetched data as direct values into Sheet1 columns G:N."""
    path = excel_path or EXCEL_PATH
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]

    ws["B2"] = datetime.now().date()

    def fmt(v):
        return round(v, 8) if v is not None else None

    for row in ws.iter_rows(min_row=5):
        ticker = row[1].value  # Col B
        if not ticker or ticker not in results:
            continue
        data = results[ticker]
        if "error" in data:
            continue
        row[6].value = data.get("price")       # G: Current Price
        row[7].value = data.get("volume")      # H: Volume
        row[8].value = fmt(data.get("1d"))     # I: 1 Day Change %
        row[9].value = fmt(data.get("5d"))     # J: 5 Day Change %
        row[10].value = fmt(data.get("1mo"))   # K: 1 Mos Change %
        row[11].value = fmt(data.get("1yr"))   # L: 1 Year Change %
        row[12].value = fmt(data.get("2yr"))   # M: 2 Year Change %
        row[13].value = fmt(data.get("5yr"))   # N: 5 Year Change %

    wb.save(path)
    wb.close()


def read_sheet1_grouped(excel_path=None, results: dict | None = None) -> dict:
    """
    Read Sheet1 structure and return holdings grouped by account, then by ETF.
    If `results` (fetched market data keyed by ticker) is provided, price/change
    data comes from there instead of Sheet1 formula cells (which Python can't evaluate).
    Returns: { "IRA": [ {etf, name, price, d1, ..., holdings: [...]}, ... ], ... }
    """
    path = excel_path or EXCEL_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]

    def _mkt(ticker):
        """Pull market data from results dict for a given ticker."""
        if not results or not ticker or ticker not in results:
            return {"price": None, "d1": None, "d5": None, "mo1": None,
                    "yr1": None, "yr2": None, "yr5": None}
        d = results[ticker]
        if "error" in d:
            return {"price": None, "d1": None, "d5": None, "mo1": None,
                    "yr1": None, "yr2": None, "yr5": None}
        return {"price": d.get("price"), "d1": d.get("1d"), "d5": d.get("5d"),
                "mo1": d.get("1mo"), "yr1": d.get("1yr"), "yr2": d.get("2yr"),
                "yr5": d.get("5yr")}

    groups: list[dict] = []
    current: dict | None = None

    for row in ws.iter_rows(min_row=5, values_only=True):
        name, ticker, weight, parent_etf, acct = row[0], row[1], row[2], row[3], row[4]

        if not name or name == "ETF":
            continue

        is_header = ticker and parent_etf and ticker == parent_etf

        if is_header:
            if current:
                groups.append(current)
            current = {
                "etf": ticker,
                "name": str(name).replace("\xa0", " ").strip(),
                "acct": acct or "",
                **_mkt(ticker),
                "holdings": [],
                "_acct_confirmed": False,
            }
        elif current and name:
            if acct and not current["_acct_confirmed"]:
                current["acct"] = acct
                current["_acct_confirmed"] = True
            current["holdings"].append({
                "name": str(name).replace("\xa0", " ").strip(),
                "ticker": ticker or "",
                "weight": weight,
                **_mkt(ticker),
            })

    if current:
        groups.append(current)

    wb.close()

    by_account: dict[str, list] = {}
    for g in groups:
        g.pop("_acct_confirmed", None)
        by_account.setdefault(g["acct"] or "Unknown", []).append(g)

    return by_account


def write_to_excel(results: dict, excel_path=None):
    """Write fetched results into Daily Update and flag Excel to recalculate on open."""
    path = excel_path or EXCEL_PATH
    wb = openpyxl.load_workbook(path)
    ws = wb["Daily Update"]

    # Tell Excel to recalculate all formulas (including Sheet1 XLOOKUPs) when opened
    if wb.calculation is None:
        from openpyxl.workbook.properties import CalcProperties
        wb.calculation = CalcProperties()
    wb.calculation.fullCalcOnLoad = True

    now = datetime.now()
    ws["B2"] = now.date()

    for row in ws.iter_rows(min_row=6):
        ticker = row[0].value
        if not ticker or ticker not in results:
            continue
        data = results[ticker]
        if "error" in data:
            continue

        def fmt(v):
            return round(v, 8) if v is not None else None

        row[2].value = data.get("price")           # C: Current Price
        row[3].value = data.get("volume")          # D: Volume
        row[4].value = fmt(data.get("1d"))         # E: 1 Day Change %
        row[5].value = fmt(data.get("5d"))         # F: 5 Day Change %
        row[6].value = fmt(data.get("1mo"))        # G: 1 Mos Change %
        row[7].value = fmt(data.get("1yr"))        # H: 1 Year Change %
        row[8].value = fmt(data.get("2yr"))        # I: 2 Year Change %
        row[9].value = fmt(data.get("5yr"))        # J: 5 Year Change %
        row[10].value = now.strftime("%Y-%m-%d %H:%M")  # K: Last Updated
        row[11].value = "fixed $1.00" if data.get("money_market") else "yfinance"  # L: Source

    wb.save(path)
    wb.close()
